import openpyxl, zipfile, time
wb = openpyxl.load_workbook('german-small.xlsx')
ws = wb.active
print('sheet name:', repr(ws.title), '| dims:', ws.dimensions)
for row in ws.iter_rows(min_row=1, max_row=4):
    print(' | '.join(f'{c.value!r}<{type(c.value).__name__},fmt={c.number_format}>' for c in row))
h = ws['A1']
print('header bold:', h.font.bold)
print('col widths:', {k: v.width for k, v in ws.column_dimensions.items()})

with zipfile.ZipFile('german-small.xlsx') as z:
    names = z.namelist()
    print('zip entries:', names)
    ss = [n for n in names if 'sharedStrings' in n]
    raw = z.read(ss[0]) if ss else z.read('xl/worksheets/sheet1.xml')
    print('xml decl:', raw[:60])
    print('umlaut bytes present (UTF-8 ü = c3bc):', b'\xc3\xbc' in raw)
    print('euro sign in header (UTF-8 e282ac):', b'\xe2\x82\xac' in raw)

t0 = time.time()
wb2 = openpyxl.load_workbook('german-100k.xlsx', read_only=True)
ws2 = wb2.active
n = sum(1 for _ in ws2.iter_rows(values_only=True))
print(f'100k file: {n} rows read by openpyxl in {time.time()-t0:.1f}s, max_row={ws2.max_row}')

print()
print('--- worksheet XML structure ---')
with zipfile.ZipFile('german-small.xlsx') as z:
    sheet = z.read('xl/worksheets/sheet1.xml').decode()
    print('has <dimension> element:', '<dimension' in sheet)
    print('sheet1.xml head:', sheet[:120])
import openpyxl.utils, subprocess, sys
print('write-excel-file version:', subprocess.run(['node','-e',
  "console.log(require('./node_modules/write-excel-file/package.json').version)"],
  capture_output=True, text=True).stdout.strip())
print('node:', subprocess.run(['node','-v'], capture_output=True, text=True).stdout.strip())
print('openpyxl:', openpyxl.__version__)
